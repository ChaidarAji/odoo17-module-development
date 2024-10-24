# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import openpyxl
import base64
from io import BytesIO

class CustomExcelUpdateData(models.Model):
    _name = 'custom.excel.update.data'
    _description = 'custom.excel.update.data'

    name = fields.Char()
    file = fields.Binary(string="File", required=True)
    filename = fields.Char(string="Filename")
    header_line_ids = fields.One2many("custom.excel.update.data.header", "update_data_id", string="Header Data")
    model_id = fields.Many2one("ir.model", string="Records Model to Update")
    error_line_ids = fields.One2many("custom.excel.update.data.error", "update_data_id", string="Error Data")
    error_line_excel_count = fields.Integer(string="Excel Line Error Count", compute="onchange_error_line_ids")

    @api.onchange('error_line_ids')
    def onchange_error_line_ids(self):
        if self.error_line_ids:
            self.error_line_excel_count = len(self.error_line_ids.filtered(lambda x: x.state == 'unresolved').mapped('excel_line'))
        else:
            self.error_line_excel_count = 0

    def generate_header(self):
        try:
            wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.file)), read_only=True)
            ws = wb.active
        except:
            raise UserError("Please insert a valid file")

        self.header_line_ids.unlink()
        header_data = self.env['custom.excel.update.data.header']
        header = True
        for record in ws.iter_rows(min_row=1, max_row=2, min_col=0,max_col=None, values_only=True):
            if header:
                for index, header_name in enumerate(record):
                    header_data.create({'name': header_name,
                                        'index_no': index,
                                        'update_data_id': self.id})
                header = False
            else:
                for header in self.header_line_ids:
                    header.write({'example': record[header.index_no]})
    
    def action_update_data(self):
        try:
            wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.file)), read_only=True)
            ws = wb.active
        except:
            raise UserError("Please insert a valid file")
        
        self.error_line_ids.unlink()
        header_domain = self.header_line_ids.filtered(lambda x: x.search_by)
        header_to_update = self.header_line_ids.filtered(lambda x: x.update_column)
        error_record = self.env['custom.excel.update.data.error']
        if not header_domain:
            raise UserError("Please select header for search record [Search By]")
        line = 1
        for record in ws.iter_rows(min_row=2, max_row=None, min_col=0,max_col=None, values_only=True):
            line += 1
            domain = []
            error_message_domain = "Record not found for "
            next_record = False
            for header in header_domain:
                if header.field_id.ttype in ['char', 'text', 'boolean', 'selection']:
                    if not record[header.index_no]:
                        error_record.create({'update_data_id': self.id,
                                             'error_message': "Search by header is empty " + header.field_id.name + ", Excel line " + str(line)})
                        next_record = True
                        break
                    domain.append((header.field_id.name, '=', record[header.index_no]))
                    error_message_domain += header.field_id.name + " = " + record[header.index_no] + ", "
                elif header.field_id.ttype == 'many2one':
                    if not record[header.index_no]:
                        error_record.create({'update_data_id': self.id,
                                             'error_message': "Search by header is empty " + header.field_id.name + ", Excel line " + str(line)})
                        next_record = True
                        break
                    domain.append((header.field_id.name+".name", '=ilike', record[header.index_no]))
                    error_message_domain += header.field_id.name + " = " + record[header.index_no] + ", "

            if next_record:
                continue
            record_id = self.env[self.model_id.model].search(domain)
            if len(record_id) > 1:
                error_message = error_message_domain + "More than one record found ["+", ".join(record_id.mapped('name'))+"], Excel line " + str(line)
                error_id = error_record.create({'update_data_id': self.id,
                                     'error_message': error_message,
                                     'data_found': "["+", ".join([str(id) for id in record_id.mapped('id')])+"]",
                                     'main_record': True,
                                     'excel_line': line})
                for value in record_id:
                    self.env['custom.excel.update.data.error.record.found'].create({'error_id': error_id.id,
                                                                                  'name': value.name,
                                                                                  'record_id': value.id,
                                                                                  'relation': self.model_id.model})
                continue
 
            if record_id:
                for header in header_to_update:
                    if record[header.index_no]:
                        if header.field_id.ttype == 'many2one':
                            update_value = self.env[header.field_id.relation].search([("name", '=ilike', record[header.index_no])])
                            if not update_value:
                                error_message = "Record not found for " + header.field_id.name + " = " + record[header.index_no]+ ", Excel line " + str(line)
                                error_record.create({'update_data_id': self.id,
                                                    'error_message': error_message,
                                                    'excel_line': line,
                                                    'main_record_id': record_id.id})
                                continue
                            else:
                                if len(update_value) > 1:
                                    error_message = "More than one record found for " + header.field_id.name + " = " + record[header.index_no]+ " ["+", ".join(update_value.mapped('name'))+"]"+ ", Excel line " + str(line)
                                    error_id = error_record.create({'update_data_id': self.id,
                                                        'error_message': error_message,
                                                        'header_line_id': header.id,
                                                        'data_found': "["+", ".join([str(id) for id in update_value.mapped('id')])+"]",
                                                        'main_record': False,
                                                        'excel_line': line,
                                                        'main_record_id': record_id.id,})
                                    for val in update_value:
                                        self.env['custom.excel.update.data.error.record.found'].create({'error_id': error_id.id,
                                                                                                      'name': val.name,
                                                                                                      'record_id': val.id,
                                                                                                      'relation': header.field_id.relation})
                                    continue
                                else:
                                    update_value = update_value.id
                        else:
                            update_value = record[header.index_no]
                    else:
                        update_value = False
                    record_id.write({header.field_id.name: update_value})
            else:
                error_record.create({'update_data_id': self.id,
                                     'error_message': error_message_domain + ", Excel line " + str(line),
                                     'excel_line': line})


class CustomExcelUpdateDataHeader(models.Model):
    _name = 'custom.excel.update.data.header'
    _description = 'custom.excel.update.data.header'

    name = fields.Char(string="Header Name")
    example = fields.Char(string="Example Value")
    index_no = fields.Integer(string="Index No.")
    update_data_id = fields.Many2one("custom.excel.update.data")
    update_column = fields.Boolean(string="Update Data", default=False)
    field_id = fields.Many2one("ir.model.fields", string="Field")
    search_by = fields.Boolean(string="Search By", default=False)

class CustomExcelUpdateDataError(models.Model):
    _name = 'custom.excel.update.data.error'
    _description = 'custom.excel.update.data.error'

    update_data_id = fields.Many2one("custom.excel.update.data")
    error_message = fields.Text(string="Error Message")
    data_found = fields.Char(string="Data Found")
    header_line_id = fields.Many2one("custom.excel.update.data.header", string="Header")
    main_record = fields.Boolean(string="Main Record", default=False)
    data_found_ids = fields.One2many("custom.excel.update.data.error.record.found", "error_id", string="Data Found")
    selection_field = fields.Many2many("custom.excel.update.data.error.record.found", "excel_update_error_record_found_rel", string="Select Record", domain="[('id', 'in', data_found_ids)]")
    excel_line = fields.Integer(string="Excel Line")
    state = fields.Selection([('unresolved', 'Unresolved'), ('resolved', 'Resolved')], string="State", default="unresolved")
    main_record_id = fields.Integer(string="Main Record ID")

    @api.onchange('selection_field')
    def onchange_selection_field(self):
        if not self.main_record:
            if len(self.selection_field)>1:
                raise UserError("Please select only one record")

    def action_manual_update(self):
        if not self.main_record:
            if len(self.selection_field)>1:
                raise UserError("Please select only one record")
            
        if not self.selection_field:
            raise UserError("Please select a record")
        else:
            try:
                wb = openpyxl.load_workbook(filename=BytesIO(base64.b64decode(self.update_data_id.file)), read_only=True)
                ws = wb.active
            except:
                raise UserError("Please insert a valid file")
            error_record = self.env['custom.excel.update.data.error']
            for record in ws.iter_rows(min_row=self.excel_line, max_row=self.excel_line, min_col=0,max_col=None, values_only=True):
                if self.main_record:
                    for header in self.update_data_id.header_line_ids:
                        if header.update_column:
                            for selection in self.selection_field:
                                record_id = self.env[selection.relation].browse(selection.record_id)
                                if record[header.index_no]:
                                    if header.field_id.ttype == 'many2one':
                                        update_value = self.env[header.field_id.relation].search([("name", '=ilike', record[header.index_no])])
                                        if not update_value:
                                            error_message = "Record not found for " + header.field_id.name + " = " + record[header.index_no]+ ", Excel line " + str(self.excel_line)
                                            error_record.create({'update_data_id': self.update_data_id.id,
                                                            'error_message': error_message,
                                                            'excel_line': self.excel_line,
                                                            'main_record_id': record_id.id})
                                            continue
                                        else:
                                            if len(update_value) > 1:
                                                error_message = "More than one record found for " + header.field_id.name + " = " + record[header.index_no]+ " ["+", ".join(update_value.mapped('name'))+"]"+ ", Excel line " + str(self.excel_line)
                                                error_id = error_record.create({'update_data_id': self.update_data_id.id,
                                                                'error_message': error_message,
                                                                'header_line_id': header.id,
                                                                'data_found': "["+", ".join([str(id) for id in update_value.mapped('id')])+"]",
                                                                'main_record': False,
                                                                'excel_line': self.excel_line,
                                                                'main_record_id': record_id.id})
                                                for value in update_value:
                                                    self.env['custom.excel.update.data.error.record.found'].create({'error_id': error_id.id,
                                                                                                                'name': value.name,
                                                                                                                'record_id': value.id,
                                                                                                                'relation': header.field_id.relation})
                                                continue
                                            else:
                                                update_value = update_value.id
                                    else:
                                        update_value = record[header.index_no]
                                else:
                                    update_value = False
                                print
                                record_id.write({header.field_id.name: update_value})
                else:
                    for selection in self.selection_field:
                        record_id = self.env[self.update_data_id.model_id.model].browse(self.main_record_id)
                        record_id.write({self.header_line_id.field_id.name: selection.record_id})
            self.state = 'resolved'
            print(self.state)

class CustomExcelUpdateDataErrorRecordFound(models.Model):
    _name = 'custom.excel.update.data.error.record.found'

    error_id = fields.Many2one("custom.excel.update.data.error")
    name = fields.Char(string="Name")
    record_id = fields.Integer(string="Record ID")
    relation = fields.Char(string="Relation")
